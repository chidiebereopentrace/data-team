"""Build mart_dev entity dictionary Excel workbook (+ dump YAML seed).

Usage:
  python data-eng/scripts/build_mart_entity_dictionary.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
MART_ROOT = ROOT / "dbt" / "models" / "mart_dev"
SCHEMA_YML = MART_ROOT / "schema.yml"
DOCS = ROOT / "docs"
OUT_XLSX = DOCS / "mart_dev_entity_dictionary.xlsx"
OUT_YAML = DOCS / "mart_entity_dictionary_seed.yaml"

sys.path.insert(0, str(Path(__file__).resolve().parent))
from mart_dictionary_columns import collect_column_rows  # noqa: E402
from mart_dictionary_data import seed_as_dict  # noqa: E402

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _autosize(ws, max_width: int = 48) -> None:
    for idx, col in enumerate(ws.columns, start=1):
        length = 0
        for cell in col[:80]:
            length = max(length, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_width, max(12, length + 2))


def _write_sheet(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_header(ws)
    _autosize(ws)


def parse_relationships() -> list[list[str]]:
    if not SCHEMA_YML.is_file():
        return []
    try:
        data = yaml.safe_load(SCHEMA_YML.read_text(encoding="utf-8")) or {}
    except Exception:
        return []
    rows: list[list[str]] = []
    for model in data.get("models") or []:
        from_table = model.get("name", "")
        for col in model.get("columns") or []:
            col_name = col.get("name", "")
            for test in col.get("tests") or []:
                if isinstance(test, dict) and "relationships" in test:
                    rel = test["relationships"] or {}
                    rows.append(
                        [
                            from_table,
                            col_name,
                            rel.get("to", ""),
                            rel.get("field", ""),
                            "schema.yml relationships",
                        ]
                    )
    return rows


def build() -> Path:
    seed = seed_as_dict()
    DOCS.mkdir(parents=True, exist_ok=True)
    OUT_YAML.write_text(
        yaml.safe_dump(seed, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )

    entities_by_name = {e["table_name"]: e for e in seed["entities"]}
    col_headers, col_rows, sql_models = collect_column_rows(
        entities=seed["entities"],
        overrides=seed.get("column_overrides") or {},
        mart_root=MART_ROOT,
    )

    entity_rows: list[list[object]] = []
    headers_e = [
        "table_name",
        "layer",
        "domain",
        "grain",
        "primary_key",
        "purpose",
        "analytical_use",
        "join_keys",
        "grain_filters",
        "source_lineage",
        "ota_report_use",
        "caveats",
    ]
    for name in sorted(set(sql_models) | set(entities_by_name)):
        meta = entities_by_name.get(name, {})
        layer = meta.get("layer") or sql_models.get(name, {}).get("layer", "")
        entity_rows.append(
            [
                name,
                layer,
                meta.get("domain", ""),
                meta.get("grain", ""),
                meta.get("primary_key", ""),
                meta.get("purpose", "See SQL model — curated description pending."),
                meta.get("analytical_use", ""),
                meta.get("join_keys", ""),
                meta.get("grain_filters", ""),
                meta.get("source_lineage", ""),
                meta.get("ota_report_use", ""),
                meta.get("caveats", ""),
            ]
        )

    rel_rows = parse_relationships()
    recipe_headers = [
        "recipe_id",
        "question",
        "tables",
        "filters",
        "join_sketch",
        "grain_warning",
        "ota_lane",
    ]
    recipe_rows = [
        [
            r["recipe_id"],
            r["question"],
            r["tables"],
            r["filters"],
            r["join_sketch"],
            r["grain_warning"],
            r["ota_lane"],
        ]
        for r in seed["analytical_recipes"]
    ]
    acf_headers = ["column", "meaning", "example"]
    acf_rows = [[r["column"], r["meaning"], r["example"]] for r in seed["acf_contract"]]

    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "Readme"
    for line in [
        ["OpenTrace mart_dev Entity Dictionary"],
        ["Dataset", "mart_dev (BigQuery gold / analytics-ready)"],
        ["Generated for", "OTA insights analysts + data consumers"],
        ["Entity count", len(entity_rows)],
        ["SQL models scanned", len(sql_models)],
        ["Seed YAML", str(OUT_YAML.relative_to(ROOT))],
        [],
        ["Sheets"],
        ["Entities", "One row per table — purpose, grain, analytical & OTA use"],
        ["Columns", "Parsed columns + curated overrides / ACF defaults"],
        ["Relationships", "FK relationships from mart_dev/schema.yml"],
        ["Analytical_Recipes", "OTA-oriented question → table recipes"],
        ["ACF_Contract", "Shared fact confidence / citation columns"],
        [],
        ["Contact", "contact@opentrace.africa"],
        ["Companion guide", "MART_DEV_OTA_ANALYST_GUIDE.docx / .md"],
    ]:
        ws_readme.append(line)
    ws_readme["A1"].font = Font(bold=True, size=14)

    ws_e = wb.create_sheet("Entities")
    _write_sheet(ws_e, headers_e, entity_rows)

    ws_c = wb.create_sheet("Columns")
    _write_sheet(ws_c, col_headers, col_rows)

    ws_r = wb.create_sheet("Relationships")
    _write_sheet(
        ws_r,
        ["from_table", "from_column", "to_ref", "to_field", "source"],
        rel_rows,
    )

    ws_a = wb.create_sheet("Analytical_Recipes")
    _write_sheet(ws_a, recipe_headers, recipe_rows)

    ws_acf = wb.create_sheet("ACF_Contract")
    _write_sheet(ws_acf, acf_headers, acf_rows)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX} ({len(entity_rows)} entities, {len(col_rows)} columns)")
    print(f"Wrote {OUT_YAML}")
    missing_seed = sorted(set(sql_models) - set(entities_by_name))
    extra_seed = sorted(set(entities_by_name) - set(sql_models))
    if missing_seed:
        print("WARNING: SQL models without curated seed:", ", ".join(missing_seed))
    if extra_seed:
        print("WARNING: Seed entities without SQL file:", ", ".join(extra_seed))
    return OUT_XLSX


if __name__ == "__main__":
    build()
