"""Build OpenTrace_Mart_Entity_Dictionary.xlsx (branded workbook).

Usage:
  python data-eng/scripts/build_opentrace_mart_entity_dictionary.py
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "OpenTrace_Mart_Entity_Dictionary.xlsx"

sys.path.insert(0, str(Path(__file__).resolve().parent))
from mart_dictionary_columns import collect_column_rows  # noqa: E402
from mart_dictionary_data import ENTITIES  # noqa: E402

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# table_name -> indicator class tags
INDICATOR_TAGS: dict[str, str] = {
    "fct_employment": "EL",
    "fct_economics": "EL,FVC",
    "fct_household": "EL,GYI,FS",
    "fct_gender_inclusion": "GYI",
    "fct_food_balance": "FVC,FS",
    "fct_prices": "PRC,FVC",
    "fct_trade": "FVC",
    "fct_forestry": "FVC,PROD",
    "fct_food_security": "FS",
    "fct_humanitarian": "FS",
    "fct_production": "PROD",
    "fct_yield": "PROD",
    "agg_production_country_year": "PROD",
    "agg_production_country_season": "PROD",
    "agg_prices_country_month": "PRC",
    "agg_food_security_country_month": "FS",
    "agg_food_balance_country_year": "FVC,FS",
    "fct_climate": "CLIM",
    "fct_air_quality": "CLIM",
    "fct_soil_health": "SOIL",
    "agg_soil_admin": "SOIL",
    "fct_land_inputs": "SOIL,INP",
    "fct_machinery": "INP",
    "fct_animal_health": "AH",
    "fct_food_hazards": "AH",
    "fct_insurance": "AH",
    "fct_vegetation": "VEG",
    "fct_emissions": "ENV",
    "agg_emissions_country_year": "ENV",
    "fct_hdi": "HDI",
    "agg_hdi_latest": "HDI",
    "agg_economics_country_year": "EL",
    "agg_employment_country_year": "EL,GYI",
    "agg_forestry_country_year": "FVC,PROD",
    "fct_biodiversity": "BIO",
    "fct_protected_areas": "BIO",
    "fct_germplasm": "BIO",
    "fct_investment": "RES",
    "dim_research_project": "RES",
    "dim_organisation": "RES",
    "dim_person": "RES",
    "bridge_research_org": "RES",
    "bridge_research_person": "RES",
    "bridge_geography_aez": "SOIL,CLIM",
    "dim_geography": "All",
    "dim_source": "All",
    "dim_date": "All",
    "dim_product": "PROD,PRC,FVC",
    "dim_season": "PROD",
    "dim_classification": "FS",
    "dim_market": "PRC",
    "dim_ref_country": "All",
}

INDICATOR_CLASS_MAP: list[tuple[str, str, str, str, str]] = [
    (
        "EL",
        "Economic & Livelihood",
        "fct_employment, fct_economics, fct_household",
        "unit; measurement_form; survey vs official",
        "ILRI income as national GDP",
    ),
    (
        "GYI",
        "Gender, Youth & Inclusion",
        "fct_gender_inclusion, fct_household, fct_employment by sex",
        "national SDG vs household sample",
        "HH control = SDG 5.a.1",
    ),
    (
        "FVC",
        "Food System & Value Chain",
        "fct_economics VA, fct_food_balance, fct_prices, fct_trade, fct_forestry",
        "measurement_form; trade_grain; price source_key",
        "SUA losses as measured farm-gate waste; blend FEWS border with FAOSTAT year",
    ),
    (
        "FS",
        "Food Security & Nutrition",
        "fct_food_security, fct_household FIES/energy, fct_humanitarian, FBS food",
        "measure_type; scenario_name",
        "Child stunting/wasting; mix population with classification",
    ),
    (
        "PROD",
        "Agricultural Production",
        "fct_production, fct_yield, agg_production_*",
        "production_grain on fct_production; keep yield SoT separate",
        "Blend FAOSTAT country-year with FNID-season yield",
    ),
    (
        "PRC",
        "Prices & Markets",
        "fct_prices, agg_prices_country_month",
        "source_key; price_type; month 1-12",
        "WFP common_unit_price; month codes 70xx as calendar",
    ),
    (
        "CLIM",
        "Climate & Weather",
        "fct_climate, fct_air_quality",
        "climate_grain",
        "Mix NASA point with ClimateWatch country model as one series",
    ),
    (
        "SOIL",
        "Soil Health & Land",
        "fct_soil_health, agg_soil_admin, fct_land_inputs",
        "source_key; input_grain",
        "Average iSDA with ISRIC; regen adoption rates",
    ),
    (
        "AH",
        "Animal Health",
        "fct_animal_health, fct_food_hazards, fct_insurance",
        "farm vs study grain",
        "Study prevalence as farm CDS",
    ),
    (
        "VEG",
        "Vegetation",
        "fct_vegetation",
        "vegetation_grain",
        "NDVI as yield",
    ),
    (
        "ENV",
        "Environment / emissions",
        "fct_emissions, agg_emissions_country_year",
        "element; source_key",
        "Intensity as total",
    ),
    (
        "INP",
        "Inputs",
        "fct_land_inputs, fct_machinery",
        "input_grain use vs trade",
        "Trade partner rows as domestic use",
    ),
    (
        "HDI",
        "Human Development",
        "fct_hdi, agg_hdi_latest, GDP in fct_economics",
        "country-year; geo must resolve",
        "HDI without geography_key / country_iso3",
    ),
    (
        "BIO",
        "Biodiversity & protection",
        "fct_biodiversity, fct_protected_areas, fct_germplasm",
        "point/feature grain; as_of_date_basis",
        "Occurrence as national biodiversity index",
    ),
    (
        "RES",
        "Research system",
        "fct_investment, dim_research_project, bridges research_*",
        "org_source; ASTI vs OpenAIRE",
        "Project counts as national R&D spend without source_key",
    ),
]

LAYER_LABEL = {"dim": "dimension", "fact": "fact", "agg": "aggregate", "bridge": "bridge"}


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _autosize(ws, max_width: int = 56) -> None:
    for idx, col in enumerate(ws.columns, start=1):
        length = 0
        for cell in col[:100]:
            length = max(length, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_width, max(12, length + 2))


def _write(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_header(ws)
    _autosize(ws)


def build() -> Path:
    wb = Workbook()

    # 00_README
    ws0 = wb.active
    ws0.title = "00_README"
    for line in [
        ["OpenTrace mart_dev — entity dictionary"],
        [],
        [
            "Companion to OpenTrace_Mart_Complete_Guide.md (ERD + indicator-class writing guide).\n"
            "Physical dataset: mart_dev. Every fact must carry source_key and as_of_date.\n"
            "Sparse nulls on union facts are often expected — score completeness by grain x source."
        ],
        [],
        ["Version", "Aug 2026 — post production/yield split"],
        ["Live models", "63 (22 dim + 27 fact + 11 agg + 3 bridge)"],
        ["Regenerate", "python scripts/build_opentrace_mart_entity_dictionary.py"],
        [],
        ["Related"],
        ["OpenTrace_Mart_Complete_Guide.md", "Indicator-class + ERD guide"],
        ["mart_dev_entity_dictionary.xlsx", "Engineer dump (Relationships / Recipes / ACF)"],
        ["MART_DEV_OTA_ANALYST_GUIDE.docx", "OTA report playbook"],
        ["CATALOG_TO_MART_MAP.md", "Grain rules + ACF contract"],
        [],
        ["Sheets"],
        ["01_Table_Catalogue", "All tables: grain, PK, joins, indicator classes"],
        ["02_Common_Columns", "Shared fact / ACF columns (cheat-sheet)"],
        ["03_Indicator_Class_Map", "EL…RES writing map"],
        ["04_Insight_Template", "Claim template + worked examples"],
        [
            "05_Columns",
            "Full per-table columns + curated descriptions (SQL + COLUMN_OVERRIDES / ACF)",
        ],
    ]:
        ws0.append(line)
    ws0["A1"].font = Font(bold=True, size=14)

    # 01_Table_Catalogue
    cat_rows: list[list[object]] = []
    layer_order = {"dim": 0, "fact": 1, "agg": 2, "bridge": 3}
    for e in sorted(ENTITIES, key=lambda x: (layer_order.get(x["layer"], 9), x["table_name"])):
        name = e["table_name"]
        cat_rows.append(
            [
                LAYER_LABEL.get(e["layer"], e["layer"]),
                name,
                e.get("purpose", ""),
                e.get("grain", ""),
                e.get("primary_key", ""),
                e.get("join_keys", ""),
                INDICATOR_TAGS.get(name, ""),
            ]
        )
    ws1 = wb.create_sheet("01_Table_Catalogue")
    _write(
        ws1,
        [
            "Layer",
            "Table",
            "Entity description",
            "Grain",
            "Primary keys",
            "Joins to",
            "Indicator classes",
        ],
        cat_rows,
    )

    # 02_Common_Columns
    common = [
        ["All facts", "source_key", "STRING", "FK to dim_source", "Required citation / ACF source_id"],
        ["All facts", "source_natural_key", "STRING", "Staging lineage string", "Debug; must exist in dim_source"],
        ["All facts", "as_of_date", "DATE", "Observation or pipeline date", "ACF freshness; filter with ranges"],
        ["All facts", "as_of_date_basis", "STRING", "observation | loaded_at", "Flag pipeline-time dates"],
        ["All facts", "tier", "INT", "ACF producer scale 1/2/3", "From source registry / fact"],
        ["All facts", "data_level", "STRING", "national|sub_national|community|point", "Warehouse row resolution"],
        ["All facts", "geo_scope", "STRING", "Same bands; null if geography_key null", "ACF place band on facts"],
        ["All facts", "place_scope", "ARRAY", "Comparable place labels", "ACF geo overlap"],
        ["All facts", "metric", "STRING", "Stable measure slug", "ACF metric identity"],
        ["All facts", "source_id", "STRING", "Usually = source_key", "ACF source_id"],
        ["All facts", "value", "FLOAT", "Primary numeric measure", "Cite with unit"],
        ["All facts", "unit", "STRING", "Unit of value", "Never invent"],
        ["All facts", "geography_key", "STRING", "FK dim_geography (nullable)", "Join places"],
        ["All facts", "country_iso3", "STRING", "ISO3 when known", "Cluster / filter"],
        ["All facts", "loaded_at", "TIMESTAMP", "Pipeline load time", "Not observation unless basis says so"],
        ["fct_production", "production_grain", "STRING", "physical|index|gross_value", "Filter before area/qty"],
        ["fct_yield", "season_key", "STRING", "FK dim_season", "FNID-season path only"],
        ["fct_trade", "trade_grain", "STRING", "faostat_country_year|fews_border_month", "Filter before blend"],
        ["fct_climate", "climate_grain", "STRING", "point_obs|country_model", "Do not mix series"],
        ["fct_food_security", "measure_type", "STRING", "classification|population|…", "pct_phase* on population"],
        ["dim_geography", "data_level", "STRING", "national|sub_national|community|point", "Dim resolution (not fact geo_scope)"],
        ["dim_geography", "geo_level", "STRING", "country|admin1|admin2|city|fnid|…", "Place type"],
    ]
    ws2 = wb.create_sheet("02_Common_Columns")
    _write(ws2, ["Applies to", "Column", "Type", "Description", "Analytical / ACF use"], common)

    # 03_Indicator_Class_Map
    ws3 = wb.create_sheet("03_Indicator_Class_Map")
    _write(
        ws3,
        ["Class ID", "Class name", "Write insights from", "Always filter", "Do not claim"],
        [list(r) for r in INDICATOR_CLASS_MAP],
    )

    # 04_Insight_Template
    ws4 = wb.create_sheet("04_Insight_Template")
    ws4.append(["Insight writing template"])
    ws4.append([])
    ws4.append(
        [
            "Claim: <metric> <direction/magnitude> in <place> during <period>\n"
            "Grain: country-year | fnid-season | household | point\n"
            "Source: organisation + source_natural_key + tier\n"
            "Unit: as stored / converted\n"
            "Caveats: grain, sample vs official, expected nulls, as_of_date_basis\n"
            "ACF: tier, data_level, as_of_date, geo_scope/place_scope, metric, source_id\n\n"
            "PROD seasonal: Maize yield Ethiopia Meher 2019 = Y t/ha across FNIDs "
            "(fct_yield / yield_raw_data). Not FAOSTAT national production "
            "(fct_production, production_grain=physical).\n\n"
            "PROD national: Nigeria maize 2022 = Q tonnes "
            "(agg_production_country_year, physical grain, keep source_key).\n\n"
            "FS: FNID X month M, P% in IPC Phase 3+ (measure_type=population only).\n\n"
            "GYI: Female farm-income control = S in ILRI survey (fct_household), "
            "not fct_gender_inclusion SDG.\n\n"
            "FVC/trade: Border rice flow uses fct_trade trade_grain=fews_border_month — "
            "not FAOSTAT country-year trade."
        ]
    )
    ws4["A1"].font = Font(bold=True, size=12)
    ws4.column_dimensions["A"].width = 100
    ws4["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws4.row_dimensions[3].height = 220

    # 05_Columns — same inventory + descriptions as mart_dev_entity_dictionary.xlsx
    col_headers, col_rows, _sql_models = collect_column_rows()
    ws5 = wb.create_sheet("05_Columns")
    _write(ws5, col_headers, col_rows)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    described = sum(1 for r in col_rows if r[4])
    print(
        f"Wrote {OUT} ({len(cat_rows)} catalogue rows, "
        f"{len(col_rows)} columns, {described} with description)"
    )
    if len(cat_rows) != 63:
        print(f"WARNING: expected 63 entities, got {len(cat_rows)}")
    return OUT


if __name__ == "__main__":
    build()
